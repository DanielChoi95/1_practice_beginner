from openpyxl import Workbook
wb= Workbook()
ws= wb.active

#병합하기
ws.merge_cells("B2:D2")
ws["B2"].value = "Merged Cell"

#분할하기
# ws.unmerge_cells("B2:D2")

wb.save("sample_merge.xlsx")