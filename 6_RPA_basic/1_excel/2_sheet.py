from openpyxl import Workbook
wb = Workbook() #새 워크북 생성

ws = wb.create_sheet() #새로운 sheet를 기본 이름으로 생성
ws.title = "MySheet"
ws.sheet_properties.tabColor = "ff66ff" #RGB형태로 넣어주면 탭 색깔 변경
ws1 = wb.create_sheet("YourSheet") #주어진 이름으로 sheet 생성
#Sheet, MySheet, YourSheet
ws2 = wb.create_sheet("NewSheet", 2) #2번째 인덱스에 sheet 생성

new_ws = wb["NewSheet"] #Dict형태로 sheet에 접근

#print(wb.sheetnames) #모든 sheet이름 확인

#Sheet 복사
new_ws["A1"] = "Test"
target = wb.copy_worksheet(new_ws)
target.title = "Copied Sheet"


wb.save("sample.xlsx")
