from openpyxl import Workbook
from random import *
from openpyxl.utils.cell import coordinate_from_string

wb= Workbook()
ws= wb.active 

#1줄씩 데이터 넣기
ws.append(["번호", "영어", "수학"])
for i in range(1,11):
    ws.append([i, randint(0,100), randint(0,100)])

# col_B = ws["B"] #영어 column만 가져와서 변수에 저장
# for cell in col_B:
#     print(cell.value)

# #영어, 수학 column 함께 가지고 오기
# col_range = ws["B":"C"] 
# for cols in col_range:
#     for cell in cols:
#         print(cell.value)

# row_title= ws[1] #1번째 row만 가지고 오기
# for cell in row_title:
#     print(cell.value)

# row_range= ws[2:6] #2번째 줄에서 6번째 줄까지 가지고 오기(2,3,4,5,6)
# for rows in row_range:
#     for cell in rows:
#         print(cell.value, end=" ")
#     print()

# row_range= ws[2:ws.max_row] #2번째 줄부터 마지막 줄까지
# for rows in row_range: 
#     for cell in rows:
#         #print(cell.value, end=" ")
#         #print(cell.coordinate, end=" ") #셀의 좌표정보를 가지고 옴
#         xy= coordinate_from_string(cell.coordinate)
#         #print(xy, end=" ") #tuple 형태로 반환
#         print(xy[0], end=" ") # A
#         print(xy[1], end=" ") # 1
#     print()

#전체 rows
#print(tuple(ws.rows))
# for row in tuple(ws.rows):
#     print(row[1].value)

#전체 columns
#print(tuple(ws.columns))
# for column in tuple(ws.columns):
#     print(column[0].value)

# for row in ws.iter_rows(): #iter_rows는 데이터를 행 단위로 가져올 때
#     print(row[1].value) #1인덱스의 row값을 반환

#1번째 줄부터 5번째 줄까지, 2번째열부터 3번째열까지
for row in ws.iter_rows(min_row=1, min_col=2, max_row=5, max_col=3):
    print(row[0].value, row[1].value) #수학, 영어



wb.save("sample.xlsx")