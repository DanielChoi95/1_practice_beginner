from openpyxl import Workbook, workbook
import datetime

wb= Workbook()
ws=wb.active

ws["A1"] = datetime.datetime.today() #오늘 날짜 정보 입력
ws["A2"] = "=SUM(1,2,3)"  #합계 입력
ws["A3"] = "AVERAGE(1,2,3)"  #평균 입력

ws["A4"] = 10
ws["A5"] = 20
ws["A6"] = "SUM(A4:A5)"

wb.save("sample_formula.xlsx")