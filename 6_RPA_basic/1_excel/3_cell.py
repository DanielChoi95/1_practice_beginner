from openpyxl import Workbook
from random import *

wb= Workbook()
ws = wb.active
ws.title= "CdhSheet"


ws["A1"] = 1
ws["A2"] = 2
ws["A3"] = 3

ws["B1"] = 4
ws["B2"] = 5
ws["B3"] = 6

print(ws["A1"]) #A1 셀의 '정보'를 출력
print(ws["A1"].value) #A1 셀의 '값'을 출력, 빈칸이면 None을 출력

print(ws.cell(row=1, column=1).value) #행렬 식으로도 가능, A1
print(ws.cell(column=2, row=1).value) #B1

ws.cell(column=3, row=1, value=10) #C1셀에 10을 입력

#반복문을 이용해서 랜덤 숫자 채우기
index = 1
for x in range(1,11): #10개 row
    for y in range(1, 11): #10개 column
        #ws.cell(row=x, column=y, value=randint(0,100)) #0~100 사이의 숫자
        ws.cell(row=x, column=y, value=index)
        index +=1

wb.save("sample.xlsx")