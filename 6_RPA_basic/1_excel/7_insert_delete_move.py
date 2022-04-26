from openpyxl import load_workbook
wb=load_workbook("sample.xlsx")
ws=wb.active

#삽입
ws.insert_rows(8) #8번째 줄 위치에 새 줄 추가
ws.insert_rows(8, 5) #8번째 줄 위치에 5줄 추가

#삭제
ws.delete_rows(8) #8번째 줄 삭제
ws.delete_rows(8,3) #8번쨰 줄부터 3줄 삭제

#이동
ws.move_range("B1:C11", rows=0, cols=1) #row는 가만히, col을 한번 옮김
ws["B1"].value = "국어"

wb.save("sample_insert_rows.xlsx")