from openpyxl import load_workbook

#수식을 그대로 가져옴
# wb= load_workbook("sample_formula.xlsx")
# ws= wb.active

# for row in ws.values:
#     for cell in row:
#         print(cell)

#수식이 계산된 데이터를 가져옴
#evaluate 되지 않은 상태의 데이터는 None 이라고 표시
wb= load_workbook("sample_formula.xlsx", data_only=True)
ws= wb.active

for row in ws.values:
    for cell in row:
        print(cell)