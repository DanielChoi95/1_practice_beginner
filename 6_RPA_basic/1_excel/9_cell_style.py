from openpyxl import load_workbook
wb=load_workbook("sample.xlsx")
ws=wb.active

#번호, 영어, 수학
a1= ws["A1"] #번호
b1= ws["B1"] #영어
c1= ws["C1"] #수학

ws.column_dimensions["A"].width = 5 #A열의 넓이를 지정
ws.row_dimensions[1].height = 50 #1행의 높이를 50으로 지정

from openpyxl.styles import Font, Border, Side, PatternFill, Alignment

#스타일 적용
a1.font= Font(color= "FF0000", italic=True, bold=True) #a1셀의 글자 색 빨갛게, 이탤릭, 굵게
b1.font= Font(color= "CC33FF", name= "Arial", strike=True) #b1셀의 글자 색 핑크, "Arial"이라는 폰트 사용, 취소선 적용
c1.font= Font(color="0000FF", size=20, underline="single") #글자 크기를 20, 밑줄 적용

#테두리 적용
thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
a1.border= thin_border

#셀의 배경 색깔을 바꿈
for row in ws.rows:
    for cell in row:
        #각 셀에 대해서 정렬
        cell.alignment= Alignment(horizontal= "center", vertical= "center") #center, left, right, top, bottom

        if cell.column == 1: #A열은 제외
            continue
        
        if isinstance(cell.value, int) and cell.value > 90: #셀의 데이터 형태가 정수형태라면
            cell.fill = PatternFill(fgColor="00FF00", fill_type= "solid") #배경을 초록색으로
            cell.font = Font(color= "FF0000") 

#틀 고정
ws.freeze_panes = "B2" #B2 기준으로 틀 고정(B2의 왼쪽과 위의 정보를 스크롤 내려도 볼수있음)


wb.save("sample_style.xlsx")